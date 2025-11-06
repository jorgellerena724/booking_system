import os
import shutil
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, models
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db.models import Sum, Count, Avg, Q
from datetime import datetime, timezone
import pandas as pd
from .models import Reservation, Room
from .forms import ReservationForm, ReservationSearchForm, RoomForm
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import json
from django.conf import settings
from django.utils import timezone
from .parser_paximum import ParserPaximum
from django.db.models import Q

# Importaciones para Excel mejorado
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart.label import DataLabelList
import openpyxl.utils

from .parser_bedbankglobal import procesar_reserva_bedbankglobal

def convertir_datos_a_uppercase(form_data):
    """Convertir campos de texto a UPPER CASE para consistencia"""
    campos_uppercase = [
        'agency', 'clients_names', 'hotel', 'hotel_confirmation', 
        'meal_plan', 'valid_rates', 'nationality', 'remarks'
    ]
    
    for campo in campos_uppercase:
        if campo in form_data and form_data[campo]:
            form_data[campo] = form_data[campo].upper()
    
    return form_data

@login_required
def reservation_list(request):
    # Verificar si se solicita exportación
    if 'export' in request.GET:
        if request.GET.get('export') == 'pdf':
            return export_reservations_pdf(request)
        else:
            return export_reservations_excel(request)
    
    # Ordenamiento por fecha de creacion (descendente por defecto)
    order_by = request.GET.get('order_by', '-created_at') 
    reservations = Reservation.objects.all().order_by(order_by)
    
    # Calcular métricas para el template - MODIFICADO: PARA TODAS LAS AGENCIAS
    total_reservations = reservations.count()
    cancelled_reservations = reservations.filter(status='CXX').count()
    
    # MODIFICADO: Eliminar filtro por agencia, aplicar a TODAS las reservas
    reservations_sin_costo = reservations.filter(
        touch_cost=0
    ).exclude(status='CXX').count()
    
    reservations_sin_confirmacion = reservations.filter(
        Q(hotel_confirmation='') | Q(hotel_confirmation__isnull=True)
    ).exclude(status='CXX').count()
    
    available_reservations = total_reservations - cancelled_reservations - reservations_sin_costo - reservations_sin_confirmacion
    
    # Paginación
    paginator = Paginator(reservations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'reservations/list.html', {
        'page_obj': page_obj,
        'current_order': order_by,
        'total_reservations': total_reservations,
        'cancelled_reservations': cancelled_reservations,
        'reservations_sin_costo': reservations_sin_costo,
        'reservations_sin_confirmacion': reservations_sin_confirmacion,
        'available_reservations': available_reservations,
    })

@login_required
def reservation_search(request):
    # Verificar si se solicita exportación
    if 'export' in request.GET:
        if request.GET.get('export') == 'pdf':
            return export_reservations_pdf(request)
        else:
            return export_reservations_excel(request)
    
    form = ReservationSearchForm(request.GET or None)
    
    # Ordenamiento por fecha de check-in (ascendente por defecto)
    order_by = request.GET.get('order_by', 'date_from')
    reservations = Reservation.objects.all().order_by(order_by)
    
    search_performed = False
    
    if form.is_valid():
        search_text = form.cleaned_data.get('search_text')
        agency = form.cleaned_data.get('agency')
        hotel = form.cleaned_data.get('hotel')
        status = form.cleaned_data.get('status')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        nationality = form.cleaned_data.get('nationality')
        room_type = form.cleaned_data.get('room_type')
        hotel_confirmation = request.GET.get('hotel_confirmation')  # NUEVO FILTRO
        
        search_performed = any([search_text, agency, hotel, status, date_from, date_to, nationality, room_type, hotel_confirmation])  # Agregar hotel_confirmation
        
        if search_text:
            reservations = reservations.filter(
                models.Q(booking_code__icontains=search_text) |
                models.Q(clients_names__icontains=search_text) |
                models.Q(hotel__icontains=search_text) |
                models.Q(hotel_confirmation__icontains=search_text) |
                models.Q(valid_rates__icontains=search_text) |
                models.Q(remarks__icontains=search_text) |
                models.Q(agency__icontains=search_text) |
                models.Q(meal_plan__icontains=search_text) |
                models.Q(rooms__room_type__icontains=search_text)
            ).distinct()
        
        if agency:
            reservations = reservations.filter(agency__icontains=agency)
        if hotel:
            reservations = reservations.filter(hotel__icontains=hotel)
        if status:
            reservations = reservations.filter(status=status)
        if nationality:
            reservations = reservations.filter(nationality__icontains=nationality)
        if room_type:
            reservations = reservations.filter(rooms__room_type__icontains=room_type)
            
        if date_from:
            reservations = reservations.filter(date_from__gte=date_from)
        if date_to:
            reservations = reservations.filter(date_to__lte=date_to)
            
        # NUEVO FILTRO: Confirmación Hotel
        if hotel_confirmation == 'sin_confirmar':
            reservations = reservations.filter(
                Q(hotel_confirmation='') | Q(hotel_confirmation__isnull=True)
            )
        elif hotel_confirmation == 'confirmadas':
            reservations = reservations.exclude(
                Q(hotel_confirmation='') | Q(hotel_confirmation__isnull=True)
            )
    
    # Paginación
    paginator = Paginator(reservations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'reservations/search.html', {
        'form': form,
        'page_obj': page_obj,
        'search_performed': search_performed,
        'current_order': order_by
    })

@login_required
def export_reservations_excel(request):
    # Replicar la misma lógica de filtros que en reservation_list
    form = ReservationSearchForm(request.GET or None)
    reservations = Reservation.objects.all().order_by('date_from')
    
    if form.is_valid():
        search_text = form.cleaned_data.get('search_text')
        agency = form.cleaned_data.get('agency')
        hotel = form.cleaned_data.get('hotel')
        status = form.cleaned_data.get('status')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        nationality = form.cleaned_data.get('nationality')
        room_type = form.cleaned_data.get('room_type')  # NUEVO
        
        if search_text:
            reservations = reservations.filter(
                models.Q(booking_code__icontains=search_text) |
                models.Q(clients_names__icontains=search_text) |
                models.Q(hotel__icontains=search_text) |
                models.Q(hotel_confirmation__icontains=search_text) |
                models.Q(valid_rates__icontains=search_text) |
                models.Q(remarks__icontains=search_text) |
                models.Q(agency__icontains=search_text) |
                models.Q(meal_plan__icontains=search_text) |
                models.Q(rooms__room_type__icontains=search_text)
            ).distinct()
        
        if agency:
            reservations = reservations.filter(agency__icontains=agency)
        if hotel:
            reservations = reservations.filter(hotel__icontains=hotel)
        if status:
            reservations = reservations.filter(status=status)
        if nationality:
            reservations = reservations.filter(nationality__icontains=nationality)
        if room_type:  # NUEVO FILTRO
            reservations = reservations.filter(rooms__room_type__icontains=room_type)
            
        if date_from:
            reservations = reservations.filter(date_from__gte=date_from)
        if date_to:
            reservations = reservations.filter(date_to__lte=date_to)
    
    # Preparar datos para Excel
    data = []
    for reservation in reservations:
        data.append({
            'STATUS': reservation.get_status_display(),
            'AGENCY': reservation.agency,
            'BOOKING CODE': reservation.booking_code,
            'CLIENTS NAMES': reservation.clients_names,
            'HOTELS': reservation.hotel,
            'From': reservation.date_from.strftime('%Y-%m-%d'),
            'to': reservation.date_to.strftime('%Y-%m-%d'),
            'Hotel confirmation': reservation.hotel_confirmation,
            'TIPO HABITACIÓN': reservation.get_rooms_display(),
            'PAX ADULTOS': reservation.total_pax_ad(),
            'PAX NIÑOS': reservation.total_pax_chd(),
            'TOTAL PAX': reservation.total_pax(),
            'MEAL PLAN': reservation.meal_plan,
            'Sale Price': float(reservation.sale_price),
            'TOUCH COST': float(reservation.touch_cost),
            'PROFIT %': f"{float(reservation.profit_percentage):.2f}%" if reservation.profit_percentage else "0%",
            'VALID RATES TO APPLY': reservation.valid_rates,
            'NATIONALITY': reservation.nationality,
            'REMARKS': reservation.remarks,
        })
    
    # Crear DataFrame
    df = pd.DataFrame(data)
    
    # Crear respuesta HTTP con el Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"reservas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Escribir DataFrame a Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='RESERVAS', index=False)
        
        # Obtener el workbook y la hoja para formatear
        workbook = writer.book
        worksheet = writer.sheets['RESERVAS']
        
        # Ajustar el ancho de las columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return response

@login_required
def export_reservations_pdf(request):
    try:
        # Replicar la misma lógica de filtros que en reservation_search
        form = ReservationSearchForm(request.GET or None)
        reservations = Reservation.objects.all().order_by('date_from')
        
        # Aplicar filtros
        if form.is_valid():
            search_text = form.cleaned_data.get('search_text')
            agency = form.cleaned_data.get('agency')
            hotel = form.cleaned_data.get('hotel')
            status = form.cleaned_data.get('status')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            nationality = form.cleaned_data.get('nationality')
            room_type = form.cleaned_data.get('room_type')  # NUEVO
            
            if search_text:
                reservations = reservations.filter(
                    models.Q(booking_code__icontains=search_text) |
                    models.Q(clients_names__icontains=search_text) |
                    models.Q(hotel__icontains=search_text) |
                    models.Q(hotel_confirmation__icontains=search_text) |
                    models.Q(valid_rates__icontains=search_text) |
                    models.Q(remarks__icontains=search_text) |
                    models.Q(agency__icontains=search_text) |
                    models.Q(meal_plan__icontains=search_text) |
                    models.Q(rooms__room_type__icontains=search_text)
                ).distinct()
            
            if agency:
                reservations = reservations.filter(agency__icontains=agency)
            if hotel:
                reservations = reservations.filter(hotel__icontains=hotel)
            if status:
                reservations = reservations.filter(status=status)
            if nationality:
                reservations = reservations.filter(nationality__icontains=nationality)
            if room_type:  # NUEVO FILTRO
                reservations = reservations.filter(rooms__room_type__icontains=room_type)
                
            if date_from:
                reservations = reservations.filter(date_from__gte=date_from)
            if date_to:
                reservations = reservations.filter(date_to__lte=date_to)
        
        # Estadísticas para el reporte - EXCLUIR CANCELADAS de ingresos
        total_reservations = reservations.count()
        active_reservations = reservations.exclude(status='CXX')
        total_revenue = active_reservations.aggregate(total=Sum('sale_price'))['total'] or 0
        total_cost = active_reservations.aggregate(total=Sum('touch_cost'))['total'] or 0
        total_profit = total_revenue - total_cost
        
        # Preparar contexto
        context = {
            'reservations': reservations,
            'total_reservations': total_reservations,
            'total_revenue': total_revenue,
            'total_profit': total_profit,
            'filters_applied': any(field in request.GET for field in [
                'search_text', 'agency', 'hotel', 'status', 'date_from', 'date_to', 'nationality', 'room_type'  # Agregar room_type
            ]),
            'request': request,
            'generated_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
        }
        
        # Renderizar template
        template = get_template('reservations/report_pdf.html')
        html = template.render(context)
        
        # Crear respuesta PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f"reporte_reservas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Generar PDF
        pisa_status = pisa.CreatePDF(
            html,
            dest=response,
            encoding='UTF-8'
        )
        
        # Si hubo error al generar el PDF
        if pisa_status.err:
            return HttpResponse(f'Error al generar PDF: {pisa_status.err}')
            
        return response
        
    except Exception as e:
        return HttpResponse(f'Error inesperado al generar PDF: {str(e)}')

@login_required
def dashboard(request):
    # Filtros de fecha opcionales
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    reservations = Reservation.objects.all()
    
    # Aplicar filtros de fecha si existen
    if date_from:
        reservations = reservations.filter(date_from__gte=date_from)
    if date_to:
        reservations = reservations.filter(date_to__lte=date_to)
    
    # MÉTRICAS FINANCIERAS - EXCLUIR CANCELADAS
    active_reservations = reservations.exclude(status='CXX')
    
    total_revenue = active_reservations.aggregate(total=Sum('sale_price'))['total'] or 0
    total_cost = active_reservations.aggregate(total=Sum('touch_cost'))['total'] or 0
    total_profit = total_revenue - total_cost
    
    # Calcular margen promedio manualmente para verificar
    total_margin = 0
    count_with_margin = 0
    
    for reservation in active_reservations:
        if reservation.touch_cost and reservation.touch_cost > 0:
            individual_margin = ((reservation.sale_price - reservation.touch_cost) / reservation.touch_cost) * 100
            total_margin += individual_margin
            count_with_margin += 1
    
    avg_profit_percentage_manual = total_margin / count_with_margin if count_with_margin > 0 else 0
    avg_profit_percentage = active_reservations.aggregate(avg=Avg('profit_percentage'))['avg'] or 0
    
    # MÉTRICAS OPERATIVAS
    total_reservations = reservations.count()
    cancelled_reservations = reservations.filter(status='CXX').count()
    active_reservations_count = total_reservations - cancelled_reservations
    cancellation_rate = (cancelled_reservations / total_reservations * 100) if total_reservations > 0 else 0
    
    status_counts = reservations.values('status').annotate(count=Count('id'))
    hotel_counts = reservations.values('hotel').annotate(count=Count('id')).order_by('-count')[:10]
    agency_counts = reservations.values('agency').annotate(count=Count('id')).order_by('-count')[:10]
    
    # TENDENCIAS TEMPORALES (últimos 12 meses) - EXCLUIR CANCELADAS
    from datetime import timedelta
    from django.utils import timezone
    
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=365)
    
    monthly_data = active_reservations.filter(
        date_from__gte=start_date
    ).extra({
        'month': "strftime('%%Y-%%m', date_from)"
    }).values('month').annotate(
        count=Count('id'),
        revenue=Sum('sale_price'),
        profit=Sum('sale_price') - Sum('touch_cost')
    ).order_by('month')
    
    # Preparar datos para gráficos
    status_order = ['OK', 'PENDING', 'CXX']
    status_counts_ordered = sorted(
        status_counts, 
        key=lambda x: status_order.index(x['status']) if x['status'] in status_order else 3
    )
    
    status_data = {
        'labels': [dict(Reservation.STATUS_CHOICES).get(item['status'], item['status']) for item in status_counts_ordered],
        'data': [item['count'] for item in status_counts_ordered],
        'colors': ['#28a745', '#ffc107', '#dc3545']
    }
    
    hotel_data = {
        'labels': [item['hotel'] for item in hotel_counts],
        'data': [item['count'] for item in hotel_counts]
    }
    
    monthly_trend = {
        'labels': [item['month'] for item in monthly_data],
        'reservations': [item['count'] for item in monthly_data],
        'revenue': [float(item['revenue'] or 0) for item in monthly_data],
        'profit': [float(item['profit'] or 0) for item in monthly_data]
    }
    
    context = {
        # Métricas financieras
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'avg_profit_percentage': avg_profit_percentage,
        'avg_profit_percentage_manual': avg_profit_percentage_manual,  # Para debugging
        
        # Métricas operativas
        'total_reservations': total_reservations,
        'cancelled_reservations': cancelled_reservations,
        'active_reservations_count': active_reservations_count,
        'cancellation_rate': cancellation_rate,
        'status_counts': status_counts,
        'hotel_counts': hotel_counts[:5],
        'agency_counts': agency_counts[:5],
        
        # Datos para gráficos
        'status_data': status_data,
        'hotel_data': hotel_data,
        'monthly_trend': monthly_trend,
        
        # Filtros
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'reservations/dashboard.html', context)

@login_required
def reservation_create(request):
    if request.method == 'POST':
        # Crear copia mutable del POST
        post_data = request.POST.copy()
        
        # Convertir campos a UPPER CASE
        post_data = convertir_datos_a_uppercase(post_data)
        
        form = ReservationForm(post_data)
        
        if form.is_valid():
            try:
                # Crear la reserva pero no guardar aún
                reservation = form.save(commit=False)
                reservation.created_by = request.user
                
                # Guardar la reserva primero
                reservation.save()
                
                # Procesar las habitaciones desde el formulario
                room_data = request.POST.get('room_data')
                if room_data:
                    rooms = json.loads(room_data)
                    for room_info in rooms:
                        # Convertir tipo de habitación a UPPER CASE
                        room_type_upper = room_info['room_type'].upper()
                        Room.objects.create(
                            reservation=reservation,
                            room_type=room_type_upper,  # EN UPPER CASE
                            pax_ad=int(room_info['pax_ad']),
                            pax_chd=int(room_info['pax_chd'])
                        )
                
                messages.success(request, 'Reserva creada exitosamente!')
                return redirect('reservations:list')
                
            except Exception as e:
                messages.error(request, f'Error al crear la reserva: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ReservationForm()
    
    room_form = RoomForm()
    return render(request, 'reservations/form.html', {
        'form': form, 
        'room_form': room_form,
        'title': 'Crear Reserva'
    })

@login_required
def reservation_detail(request, pk):
    reservation = get_object_or_404(Reservation, pk=pk)
    return render(request, 'reservations/detail.html', {'reservation': reservation})

@login_required
def reservation_update(request, pk):
    reservation = get_object_or_404(Reservation, pk=pk)
    
    if request.method == 'POST':
        # Crear copia mutable del POST
        post_data = request.POST.copy()
        
        # Convertir campos a UPPER CASE
        post_data = convertir_datos_a_uppercase(post_data)
        
        form = ReservationForm(post_data, instance=reservation)
        if form.is_valid():
            # Guardar el formulario de reserva
            reservation = form.save()
            
            # Eliminar habitaciones existentes y crear nuevas
            reservation.rooms.all().delete()
            
            # Procesar las nuevas habitaciones
            room_data = request.POST.get('room_data')
            if room_data:
                rooms = json.loads(room_data)
                for room_info in rooms:
                    # Convertir tipo de habitación a UPPER CASE
                    room_type_upper = room_info['room_type'].upper()
                    Room.objects.create(
                        reservation=reservation,
                        room_type=room_type_upper,  # EN UPPER CASE
                        pax_ad=int(room_info['pax_ad']),
                        pax_chd=int(room_info['pax_chd'])
                    )
            
            messages.success(request, f'Reserva {reservation.booking_code} actualizada exitosamente!')
            return redirect('reservations:list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ReservationForm(instance=reservation)
    
    # Obtener habitaciones existentes para pre-cargar el formulario
    existing_rooms = list(reservation.rooms.values('room_type', 'pax_ad', 'pax_chd'))
    
    room_form = RoomForm()
    return render(request, 'reservations/form.html', {
        'form': form, 
        'room_form': room_form,
        'existing_rooms': existing_rooms,
        'title': 'Editar Reserva'
    })

@login_required
def reservation_delete(request, pk):
    reservation = get_object_or_404(Reservation, pk=pk)
    if request.method == 'POST':
        reservation.delete()
        messages.success(request, 'Reserva eliminada exitosamente!')
        return redirect('reservations:list')
    return render(request, 'reservations/confirm_delete.html', {'reservation': reservation})

@login_required
def financial_debug(request):
    """Vista temporal para diagnosticar problemas financieros"""
    
    # Todas las reservas activas (OK y PENDING)
    active_reservations = Reservation.objects.filter(status__in=['OK', 'PENDING'])
    
    # Reservas problemáticas con cálculo de pérdida
    negative_profit_reservations_data = []
    for reservation in active_reservations.filter(sale_price__lt=models.F('touch_cost')):
        loss = float(reservation.touch_cost) - float(reservation.sale_price)
        negative_profit_reservations_data.append({
            'reservation': reservation,
            'loss': loss
        })
    
    # Cálculos detallados
    total_revenue = active_reservations.aggregate(total=Sum('sale_price'))['total'] or 0
    total_cost = active_reservations.aggregate(total=Sum('touch_cost'))['total'] or 0
    total_profit = total_revenue - total_cost
    
    # Calcular margen manualmente y preparar datos para template
    margin_data = []
    for reservation in active_reservations:
        if reservation.touch_cost and reservation.touch_cost > 0:
            # Convertir Decimal a float para cálculos consistentes
            sale_price_float = float(reservation.sale_price)
            touch_cost_float = float(reservation.touch_cost)
            stored_margin_float = float(reservation.profit_percentage) if reservation.profit_percentage else 0.0
            
            # Calcular margen manual
            manual_margin = ((sale_price_float - touch_cost_float) / touch_cost_float) * 100
            difference = manual_margin - stored_margin_float
            
            margin_data.append({
                'reservation': reservation,
                'sale_price': sale_price_float,
                'touch_cost': touch_cost_float,
                'stored_margin': stored_margin_float,
                'manual_margin': manual_margin,
                'difference': difference
            })
    
    # Calcular margen global
    profit_margin_percentage = (float(total_profit) / float(total_cost) * 100) if total_cost and float(total_cost) > 0 else 0
    
    context = {
        'total_active_reservations': active_reservations.count(),
        'negative_profit_count': len(negative_profit_reservations_data),
        'total_revenue': float(total_revenue),
        'total_cost': float(total_cost),
        'total_profit': float(total_profit),
        'profit_margin_percentage': profit_margin_percentage,
        
        'negative_profit_reservations_data': negative_profit_reservations_data,
        'margin_data': margin_data,
        
        # Métricas del dashboard para comparar
        'dashboard_margin': active_reservations.aggregate(avg=Avg('profit_percentage'))['avg'] or 0,
    }
    
    return render(request, 'reservations/financial_debug.html', context)

@login_required
def recalculate_margins(request):
    """Vista para recalcular todos los márgenes de ganancia"""
    if request.method == 'POST':
        updated_count = 0
        reservations = Reservation.objects.all()
        
        for reservation in reservations:
            # Forzar el recálculo llamando al save()
            original_margin = reservation.profit_percentage
            reservation.save()
            
            if reservation.profit_percentage != original_margin:
                updated_count += 1
        
        messages.success(request, f'Se recalcularon {updated_count} márgenes de ganancia.')
        return redirect('reservations:financial_debug')
    
    return render(request, 'reservations/confirm_recalculate.html')


@login_required
def export_dashboard_excel(request):
    """Exportar métricas del dashboard a Excel con gráficos y formato profesional"""
    try:
        # Replicar la lógica del dashboard para obtener los datos
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        # Base queryset
        reservations = Reservation.objects.all()
        
        # Aplicar filtros de fecha si existen
        if date_from:
            reservations = reservations.filter(date_from__gte=date_from)
        if date_to:
            reservations = reservations.filter(date_to__lte=date_to)
        
        # MÉTRICAS FINANCIERAS - EXCLUIR CANCELADAS
        active_reservations = reservations.exclude(status='CXX')
        total_revenue = active_reservations.aggregate(total=Sum('sale_price'))['total'] or 0
        total_cost = active_reservations.aggregate(total=Sum('touch_cost'))['total'] or 0
        total_profit = total_revenue - total_cost
        avg_profit_percentage = active_reservations.aggregate(avg=Avg('profit_percentage'))['avg'] or 0
        
        # MÉTRICAS OPERATIVAS
        total_reservations = reservations.count()
        cancelled_reservations = reservations.filter(status='CXX').count()
        active_reservations_count = total_reservations - cancelled_reservations
        cancellation_rate = (cancelled_reservations / total_reservations * 100) if total_reservations > 0 else 0
        
        status_counts = reservations.values('status').annotate(count=Count('id'))
        hotel_counts = reservations.values('hotel').annotate(count=Count('id')).order_by('-count')[:10]
        agency_counts = reservations.values('agency').annotate(count=Count('id')).order_by('-count')[:10]
        
        # TENDENCIAS MENSUALES
        from datetime import timedelta
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=365)
        
        monthly_data = active_reservations.filter(
            date_from__gte=start_date
        ).extra({
            'month': "strftime('%%Y-%%m', date_from)"
        }).values('month').annotate(
            count=Count('id'),
            revenue=Sum('sale_price')
        ).order_by('month')
        
        # Crear Workbook
        wb = Workbook()
        
        # ===== HOJA 1: DASHBOARD VISUAL =====
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        metric_font = Font(bold=True, size=12)
        metric_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Título principal
        ws_dashboard.merge_cells('A1:H1')
        ws_dashboard['A1'] = "DASHBOARD DE RESERVAS - REPORTE COMPLETO"
        ws_dashboard['A1'].font = Font(bold=True, size=16, color="366092")
        ws_dashboard['A1'].alignment = Alignment(horizontal='center')
        
        # Información de filtros
        if date_from or date_to:
            filter_text = f"Filtros aplicados: "
            if date_from:
                filter_text += f"Desde {date_from} "
            if date_to:
                filter_text += f"Hasta {date_to}"
            ws_dashboard['A2'] = filter_text
            ws_dashboard['A2'].font = Font(italic=True, color="666666")
        
        # MÉTRICAS PRINCIPALES - Fila 4
        metrics_row = 4
        metrics = [
            ("Total Reservas", total_reservations, "A"),
            ("Reservas Activas", active_reservations_count, "B"),
            ("Canceladas", cancelled_reservations, "C"),
            ("Tasa Cancelación", f"{cancellation_rate:.1f}%", "D"),
            ("Ingresos Totales", f"${total_revenue:,.2f}", "E"),
            ("Costo Total", f"${total_cost:,.2f}", "F"),
            ("Ganancia Neta", f"${total_profit:,.2f}", "G"),
            ("Margen Promedio", f"{avg_profit_percentage:.2f}%", "H")
        ]
        
        for title, value, col in metrics:
            cell = f"{col}{metrics_row}"
            ws_dashboard[f"{col}{metrics_row-1}"] = title
            ws_dashboard[f"{col}{metrics_row-1}"].font = metric_font
            ws_dashboard[f"{col}{metrics_row-1}"].fill = header_fill
            ws_dashboard[f"{col}{metrics_row-1}"].alignment = Alignment(horizontal='center')
            ws_dashboard[f"{col}{metrics_row-1}"].border = border
            
            ws_dashboard[cell] = value
            ws_dashboard[cell].font = Font(size=11, bold=True)
            ws_dashboard[cell].alignment = Alignment(horizontal='center')
            ws_dashboard[cell].border = border
        
        # RESERVAS POR ESTADO - Fila 7
        status_row = 7
        ws_dashboard[f'A{status_row}'] = "RESERVAS POR ESTADO"
        ws_dashboard[f'A{status_row}'].font = header_font
        ws_dashboard[f'A{status_row}'].fill = header_fill
        ws_dashboard.merge_cells(f'A{status_row}:B{status_row}')
        
        current_row = status_row + 1
        ws_dashboard[f'A{current_row}'] = "Estado"
        ws_dashboard[f'B{current_row}'] = "Cantidad"
        ws_dashboard[f'A{current_row}'].font = metric_font
        ws_dashboard[f'B{current_row}'].font = metric_font
        
        current_row += 1
        status_data_start = current_row
        
        for status in status_counts:
            status_name = dict(Reservation.STATUS_CHOICES).get(status['status'], status['status'])
            ws_dashboard[f'A{current_row}'] = status_name
            ws_dashboard[f'B{current_row}'] = status['count']
            current_row += 1
        
        # GRÁFICO DE PASTEL - RESERVAS POR ESTADO
        pie_chart = PieChart()
        labels = Reference(ws_dashboard, min_col=1, min_row=status_data_start, max_row=current_row-1)
        data = Reference(ws_dashboard, min_col=2, min_row=status_data_start-1, max_row=current_row-1)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        pie_chart.title = "Reservas por Estado"
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        ws_dashboard.add_chart(pie_chart, "D7")
        
        # TOP 5 HOTELES - Fila 7, Columna G
        hotels_row = 7
        ws_dashboard[f'G{hotels_row}'] = "TOP 5 HOTELES"
        ws_dashboard[f'G{hotels_row}'].font = header_font
        ws_dashboard[f'G{hotels_row}'].fill = header_fill
        ws_dashboard.merge_cells(f'G{hotels_row}:H{hotels_row}')
        
        current_hotel_row = hotels_row + 1
        ws_dashboard[f'G{current_hotel_row}'] = "Hotel"
        ws_dashboard[f'H{current_hotel_row}'] = "Reservas"
        ws_dashboard[f'G{current_hotel_row}'].font = metric_font
        ws_dashboard[f'H{current_hotel_row}'].font = metric_font
        
        current_hotel_row += 1
        hotel_data_start = current_hotel_row
        
        for hotel in hotel_counts[:5]:
            ws_dashboard[f'G{current_hotel_row}'] = hotel['hotel'][:30]  # Truncar nombres largos
            ws_dashboard[f'H{current_hotel_row}'] = hotel['count']
            current_hotel_row += 1
        
        # GRÁFICO DE BARRAS - TOP HOTELES
        bar_chart_hotels = BarChart()
        bar_chart_hotels.type = "col"
        bar_chart_hotels.style = 10
        bar_chart_hotels.title = "Top 5 Hoteles"
        bar_chart_hotels.y_axis.title = 'Reservas'
        bar_chart_hotels.x_axis.title = 'Hoteles'
        
        labels = Reference(ws_dashboard, min_col=7, min_row=hotel_data_start, max_row=current_hotel_row-1)
        data = Reference(ws_dashboard, min_col=8, min_row=hotel_data_start-1, max_row=current_hotel_row-1)
        bar_chart_hotels.add_data(data, titles_from_data=True)
        bar_chart_hotels.set_categories(labels)
        ws_dashboard.add_chart(bar_chart_hotels, "J7")
        
        # TOP 5 AGENCIAS - Fila 20
        agencies_row = 20
        ws_dashboard[f'A{agencies_row}'] = "TOP 5 AGENCIAS"
        ws_dashboard[f'A{agencies_row}'].font = header_font
        ws_dashboard[f'A{agencies_row}'].fill = header_fill
        ws_dashboard.merge_cells(f'A{agencies_row}:B{agencies_row}')
        
        current_agency_row = agencies_row + 1
        ws_dashboard[f'A{current_agency_row}'] = "Agencia"
        ws_dashboard[f'B{current_agency_row}'] = "Reservas"
        ws_dashboard[f'A{current_agency_row}'].font = metric_font
        ws_dashboard[f'B{current_agency_row}'].font = metric_font
        
        current_agency_row += 1
        agency_data_start = current_agency_row
        
        for agency in agency_counts[:5]:
            ws_dashboard[f'A{current_agency_row}'] = agency['agency']
            ws_dashboard[f'B{current_agency_row}'] = agency['count']
            current_agency_row += 1
        
        # GRÁFICO DE BARRAS - TOP AGENCIAS
        bar_chart_agencies = BarChart()
        bar_chart_agencies.type = "col"
        bar_chart_agencies.style = 10
        bar_chart_agencies.title = "Top 5 Agencias"
        bar_chart_agencies.y_axis.title = 'Reservas'
        bar_chart_agencies.x_axis.title = 'Agencias'
        
        labels = Reference(ws_dashboard, min_col=1, min_row=agency_data_start, max_row=current_agency_row-1)
        data = Reference(ws_dashboard, min_col=2, min_row=agency_data_start-1, max_row=current_agency_row-1)
        bar_chart_agencies.add_data(data, titles_from_data=True)
        bar_chart_agencies.set_categories(labels)
        ws_dashboard.add_chart(bar_chart_agencies, "D20")
        
        # TENDENCIA MENSUAL - Fila 20, Columna G
        trend_row = 20
        ws_dashboard[f'G{trend_row}'] = "TENDENCIA MENSUAL (Últimos 12 meses)"
        ws_dashboard[f'G{trend_row}'].font = header_font
        ws_dashboard[f'G{trend_row}'].fill = header_fill
        ws_dashboard.merge_cells(f'G{trend_row}:H{trend_row}')
        
        current_trend_row = trend_row + 1
        ws_dashboard[f'G{current_trend_row}'] = "Mes"
        ws_dashboard[f'H{current_trend_row}'] = "Reservas"
        ws_dashboard[f'I{current_trend_row}'] = "Ingresos"
        ws_dashboard[f'G{current_trend_row}'].font = metric_font
        ws_dashboard[f'H{current_trend_row}'].font = metric_font
        ws_dashboard[f'I{current_trend_row}'].font = metric_font
        
        current_trend_row += 1
        trend_data_start = current_trend_row
        
        for monthly in monthly_data:
            ws_dashboard[f'G{current_trend_row}'] = monthly['month']
            ws_dashboard[f'H{current_trend_row}'] = monthly['count']
            ws_dashboard[f'I{current_trend_row}'] = float(monthly['revenue'] or 0)
            current_trend_row += 1
        
        # GRÁFICO DE LÍNEAS - TENDENCIA
        line_chart = LineChart()
        line_chart.title = "Tendencia Mensual"
        line_chart.style = 12
        line_chart.y_axis.title = 'Reservas'
        line_chart.x_axis.title = 'Meses'
        
        # Datos de reservas
        labels = Reference(ws_dashboard, min_col=7, min_row=trend_data_start, max_row=current_trend_row-1)
        data = Reference(ws_dashboard, min_col=8, min_row=trend_data_start-1, max_row=current_trend_row-1)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(labels)
        ws_dashboard.add_chart(line_chart, "J20")
        
        # Ajustar anchos de columnas
        column_widths = {'A': 25, 'B': 15, 'C': 15, 'D': 15, 'E': 18, 'F': 18, 'G': 20, 'H': 15, 'I': 15}
        for col, width in column_widths.items():
            ws_dashboard.column_dimensions[col].width = width
        
        # ===== HOJA 2: DATOS DETALLADOS =====
        ws_detalles = wb.create_sheet("Datos Detallados")
        
        # Encabezados
        headers = [
            "Estado", "Agencia", "Booking Code", "Clientes", "Hotel", 
            "Check-in", "Check-out", "Precio Venta", "Costo", "Margen %",
            "Total PAX", "Plan Comidas", "Nacionalidad"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_detalles.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Datos de reservas
        row_num = 2
        for reservation in reservations:
            ws_detalles.cell(row=row_num, column=1, value=reservation.get_status_display())
            ws_detalles.cell(row=row_num, column=2, value=reservation.agency)
            ws_detalles.cell(row=row_num, column=3, value=reservation.booking_code)
            ws_detalles.cell(row=row_num, column=4, value=reservation.clients_names)
            ws_detalles.cell(row=row_num, column=5, value=reservation.hotel)
            ws_detalles.cell(row=row_num, column=6, value=reservation.date_from.strftime('%Y-%m-%d'))
            ws_detalles.cell(row=row_num, column=7, value=reservation.date_to.strftime('%Y-%m-%d'))
            ws_detalles.cell(row=row_num, column=8, value=float(reservation.sale_price))
            ws_detalles.cell(row=row_num, column=9, value=float(reservation.touch_cost))
            ws_detalles.cell(row=row_num, column=10, value=float(reservation.profit_percentage or 0))
            ws_detalles.cell(row=row_num, column=11, value=reservation.total_pax())
            ws_detalles.cell(row=row_num, column=12, value=reservation.meal_plan)
            ws_detalles.cell(row=row_num, column=13, value=reservation.nationality)
            row_num += 1
        
        # Ajustar anchos de columnas
        for column in ws_detalles.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_detalles.column_dimensions[column_letter].width = adjusted_width
        
        # ===== HOJA 3: MÉTRICAS FINANCIERAS =====
        ws_finanzas = wb.create_sheet("Métricas Financieras")
        
        # Encabezado
        ws_finanzas['A1'] = "ANÁLISIS FINANCIERO DETALLADO"
        ws_finanzas['A1'].font = Font(bold=True, size=14, color="366092")
        ws_finanzas.merge_cells('A1:C1')
        
        # Métricas financieras
        finanzas_data = [
            ("INGRESOS TOTALES", f"${total_revenue:,.2f}"),
            ("COSTOS TOTALES", f"${total_cost:,.2f}"),
            ("GANANCIA NETA", f"${total_profit:,.2f}"),
            ("MARGEN NETO", f"{(total_profit/total_revenue*100) if total_revenue > 0 else 0:.2f}%"),
            ("MARGEN PROMEDIO", f"{avg_profit_percentage:.2f}%"),
            ("TASA DE CANCELACIÓN", f"{cancellation_rate:.2f}%"),
            ("RESERVAS ACTIVAS", active_reservations_count),
            ("VALOR PROMEDIO POR RESERVA", f"${(total_revenue/active_reservations_count) if active_reservations_count > 0 else 0:.2f}")
        ]
        
        for idx, (title, value) in enumerate(finanzas_data, start=3):
            ws_finanzas[f'A{idx}'] = title
            ws_finanzas[f'A{idx}'].font = Font(bold=True)
            ws_finanzas[f'B{idx}'] = value
            ws_finanzas[f'B{idx}'].font = Font(size=12)
        
        # Ajustar anchos
        ws_finanzas.column_dimensions['A'].width = 30
        ws_finanzas.column_dimensions['B'].width = 20
        
        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"dashboard_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return HttpResponse(f'Error al generar Excel: {str(e)}\n\nDetalles: {error_details}')
    
@login_required
def backup_management(request):
    """Página principal de gestión de backups"""
    try:
        # Información de la base de datos actual
        db_path = settings.DATABASES['default']['NAME']
        db_size = "N/A"
        db_modified = "N/A"
        
        if os.path.exists(db_path):
            db_size = f"{os.path.getsize(db_path) / 1024 / 1024:.2f} MB"
            mod_time = os.path.getmtime(db_path)
            db_modified = timezone.datetime.fromtimestamp(mod_time).strftime('%d/%m/%Y %H:%M:%S')
        
        context = {
            'db_path': db_path,
            'db_size': db_size,
            'db_modified': db_modified,
        }
        return render(request, 'reservations/backup_management.html', context)
    
    except Exception as e:
        messages.error(request, f'Error al cargar información de la base de datos: {str(e)}')
        return redirect('reservations:list')

@login_required
def backup_download(request):
    """Descargar backup de la base de datos"""
    try:
        db_path = settings.DATABASES['default']['NAME']
        
        if not os.path.exists(db_path):
            messages.error(request, 'No se encontró la base de datos.')
            return redirect('reservations:backup_management')
        
        # Crear nombre del archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'backup_reservas_{timestamp}.sqlite3'
        
        # Crear respuesta con el archivo
        with open(db_path, 'rb') as db_file:
            response = HttpResponse(db_file.read(), content_type='application/x-sqlite3')
            response['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
            response['Content-Length'] = os.path.getsize(db_path)
        
        messages.success(request, f'Backup descargado exitosamente: {backup_filename}')
        return response
        
    except Exception as e:
        messages.error(request, f'Error al crear backup: {str(e)}')
        return redirect('reservations:backup_management')

@login_required
def backup_restore(request):
    """Restaurar base de datos desde backup"""
    if request.method == 'POST':
        try:
            # Verificar que se haya subido un archivo
            if 'backup_file' not in request.FILES:
                messages.error(request, 'No se seleccionó ningún archivo.')
                return redirect('reservations:backup_management')
            
            backup_file = request.FILES['backup_file']
            
            # Validaciones del archivo
            if not backup_file.name.endswith('.sqlite3'):
                messages.error(request, 'El archivo debe ser una base de datos SQLite (.sqlite3).')
                return redirect('reservations:backup_management')
            
            if backup_file.size == 0:
                messages.error(request, 'El archivo está vacío.')
                return redirect('reservations:backup_management')
            
            # Límite de tamaño: 100MB
            if backup_file.size > 100 * 1024 * 1024:
                messages.error(request, 'El archivo es demasiado grande (máximo 100MB).')
                return redirect('reservations:backup_management')
            
            db_path = settings.DATABASES['default']['NAME']
            
            # Crear backup de la base de datos actual
            if os.path.exists(db_path):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_current = f"{db_path}.backup_{timestamp}"
                shutil.copy2(db_path, backup_current)
                messages.info(request, f'Backup automático creado: {os.path.basename(backup_current)}')
            
            # Guardar el archivo subido
            with open(db_path, 'wb') as destination:
                for chunk in backup_file.chunks():
                    destination.write(chunk)
            
            # Verificar que la nueva base de datos sea válida
            try:
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' LIMIT 1;")
            except Exception as e:
                # Restaurar el backup automático si la nueva BD es inválida
                if 'backup_current' in locals() and os.path.exists(backup_current):
                    shutil.copy2(backup_current, db_path)
                    messages.error(request, f'El archivo no es una base de datos SQLite válida. Se restauró el backup automático.')
                else:
                    messages.error(request, 'El archivo no es una base de datos SQLite válida.')
                return redirect('reservations:backup_management')
            
            messages.success(request, 'Base de datos restaurada exitosamente! Es recomendable reiniciar la aplicación.')
            return redirect('reservations:backup_management')
            
        except Exception as e:
            messages.error(request, f'Error al restaurar backup: {str(e)}')
            return redirect('reservations:backup_management')
    
    return redirect('reservations:backup_management')


@login_required
def importar_menu(request):
    """Menú principal de importación"""
    return render(request, 'reservations/importar_menu.html', {
        'titulo': 'Importar Reservas'
    })

@login_required
def importar_paximum(request):
    """Importar reservas desde archivo HTML de Paximum"""
    if request.method == 'POST' and request.FILES.get('archivo_paximum'):
        try:
            archivo = request.FILES['archivo_paximum']
            
            # Leer el archivo como bytes primero
            contenido_bytes = archivo.read()
            
            # Detectar codificación automáticamente
            try:
                import chardet
                deteccion = chardet.detect(contenido_bytes)
                encoding = deteccion['encoding'] or 'utf-8'
                confidence = deteccion['confidence']
                
                print(f"Codificación detectada: {encoding} (confianza: {confidence})")
                
                # Intentar decodificar con la codificación detectada
                contenido = contenido_bytes.decode(encoding)
                
            except ImportError:
                # Fallback si chardet no está instalado
                codificaciones = ['utf-8', 'utf-16', 'latin-1', 'cp1252', 'iso-8859-1']
                contenido = None
                
                for enc in codificaciones:
                    try:
                        contenido = contenido_bytes.decode(enc)
                        break
                    except UnicodeDecodeError:
                        continue
                
                if contenido is None:
                    # Último recurso: ignorar errores
                    contenido = contenido_bytes.decode('utf-8', errors='ignore')
            
            # Parsear el HTML
            reservas_data = ParserPaximum.parsear_html(contenido)
            reservas_creadas = 0
            
            for data_reserva in reservas_data:
                if crear_reserva_desde_paximum(data_reserva, request.user):
                    reservas_creadas += 1
            
            if reservas_creadas > 0:
                messages.success(
                    request, 
                    f'¡Importación exitosa! {reservas_creadas} reservas creadas desde Paximum.'
                )
            else:
                messages.warning(request, 'No se crearon reservas. Verifique el formato del archivo.')
                
            return redirect('reservations:list')
            
        except Exception as e:
            messages.error(request, f'Error al importar: {str(e)}')
            # Para debugging
            import traceback
            print(f"Error completo: {traceback.format_exc()}")
    
    return render(request, 'reservations/importar_paximum.html')

def crear_reserva_desde_paximum(data_reserva, user):
    """Crear una reserva en la base de datos desde datos de Paximum - VERSIÓN GENÉRICA"""
    try:
        # Verificar si ya existe una reserva con este booking_code
        if Reservation.objects.filter(booking_code=data_reserva['voucher']).exists():
            print(f"Reserva con voucher {data_reserva['voucher']} ya existe")
            return False
        
        # Concatenar nombres de huéspedes (manejar casos vacíos) - CONVERTIR A UPPER CASE
        nombres_clientes = ", ".join([
            f"{huesped.get('titulo', '')} {huesped.get('nombre', '')} {huesped.get('apellido', '')}".strip()
            for huesped in data_reserva['huespedes']
            if huesped.get('nombre') or huesped.get('apellido')
        ]).upper()  # CONVERTIR A UPPER CASE
        
        # Si no hay nombres, usar un valor por defecto
        if not nombres_clientes:
            nombres_clientes = "HUÉSPED NO ESPECIFICADO"  # EN UPPER CASE
        
        # Usar el régimen de la primera habitación o valor por defecto - CONVERTIR A UPPER CASE
        regimen = data_reserva['habitaciones'][0]['regimen'] if data_reserva['habitaciones'] else 'ALOJAMIENTO DESAYUNO'
        
        # Calcular precio total
        try:
            precio_venta = float(data_reserva['precio_total'].replace(',', '.')) if data_reserva['precio_total'] and data_reserva['precio_total'] != '0.00' else 0
        except (ValueError, AttributeError):
            precio_venta = 0
        
        # Usar nacionalidad del primer huésped o valor por defecto - CONVERTIR A UPPER CASE
        nacionalidad = data_reserva['huespedes'][0]['nacionalidad'] if data_reserva['huespedes'] and data_reserva['huespedes'][0]['nacionalidad'] else 'NO ESPECIFICADA'
        
        # Crear la reserva principal - TODOS LOS CAMPOS DE TEXTO EN UPPER CASE
        reserva = Reservation(
            status='OK',
            agency='PAXIMUM',  # UPPER CASE
            booking_code=data_reserva['voucher'],
            clients_names=nombres_clientes,  # YA EN UPPER CASE
            hotel=data_reserva['hotel'],  # YA VIENE EN UPPER CASE DEL PARSER
            date_from=data_reserva['fechas']['checkin'],
            date_to=data_reserva['fechas']['checkout'],
            meal_plan=regimen,  # YA EN UPPER CASE
            sale_price=precio_venta,
            touch_cost=0,
            nationality=nacionalidad,  # YA EN UPPER CASE
            remarks=f"IMPORTADO AUTOMÁTICAMENTE DESDE PAXIMUM. {len(data_reserva['habitaciones'])} HABITACIÓN(ES).".upper(),  # UPPER CASE
            created_by=user
        )
        reserva.save()
        
        # Crear todas las habitaciones - CONVERTIR TIPO DE HABITACIÓN A UPPER CASE
        for hab_data in data_reserva['habitaciones']:
            room = Room(
                reservation=reserva,
                room_type=hab_data['tipo'],  # YA VIENE EN UPPER CASE DEL PARSER
                pax_ad=hab_data['adultos'],
                pax_chd=hab_data['ninios']
            )
            room.save()
        
        print(f"Reserva creada exitosamente: {reserva.booking_code}")
        return True
        
    except Exception as e:
        print(f"Error creando reserva Paximum: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return False
    
def importar_bedbankglobal(request):
    """
    Vista para importar reservas desde archivos HTML de BedBankGlobal
    """
    if request.method == 'POST':
        if 'archivo_html' in request.FILES:
            archivo = request.FILES['archivo_html']
            
            try:
                # Leer el contenido del archivo
                html_bytes = archivo.read()
                file_size = len(html_bytes)
                
                print(f"Archivo recibido: {archivo.name}, tamaño: {file_size} bytes")
                
                # Detectar codificación
                encodings = ['utf-8', 'utf-16', 'latin-1', 'windows-1252']
                html_content = None
                detected_encoding = None
                
                for encoding in encodings:
                    try:
                        html_content = html_bytes.decode(encoding)
                        detected_encoding = encoding
                        print(f"Codificación detectada: {encoding}")
                        break
                    except UnicodeDecodeError:
                        continue
                
                if html_content is None:
                    messages.error(request, "No se pudo decodificar el archivo. Codificación no soportada.")
                    return redirect('reservations:importar_bedbankglobal')
                
                print(f"Contenido leído: {len(html_content)} caracteres")
                
                # Procesar el archivo con el parser mejorado
                from .parser_bedbankglobal import procesar_reserva_bedbankglobal
                
                reservas = procesar_reserva_bedbankglobal(html_content)
                
                if not reservas:
                    messages.warning(request, "No se encontraron reservas en el archivo. Verifique el formato.")
                    return redirect('reservations:importar_bedbankglobal')
                
                reservas_creadas = 0
                reservas_con_error = 0
                
                for i, reserva_data in enumerate(reservas, 1):
                    print(f"Procesando reserva {i}: {reserva_data.get('booking_code', 'N/A')}")
                    
                    # Validar datos críticos
                    if not reserva_data.get('fecha_entrada') or not reserva_data.get('fecha_salida'):
                        print(f"❌ Error: Fechas faltantes para reserva {reserva_data.get('booking_code', 'N/A')}")
                        reservas_con_error += 1
                        messages.warning(request, f"Reserva {reserva_data.get('booking_code', 'N/A')}: Faltan fechas de check-in/out")
                        continue
                    
                    if not reserva_data.get('pasajeros'):
                        print(f"❌ Error: No hay pasajeros para reserva {reserva_data.get('booking_code', 'N/A')}")
                        reservas_con_error += 1
                        messages.warning(request, f"Reserva {reserva_data.get('booking_code', 'N/A')}: No se encontraron pasajeros")
                        continue
                    
                    try:
                        # ADAPTAR DATOS al formato que espera crear_reserva_desde_bedbankglobal
                        datos_adaptados = {
                            'voucher': reserva_data['booking_code'],
                            'hotel': reserva_data.get('hotel', '').upper(),
                            'fechas': {
                                'checkin': reserva_data['fecha_entrada'],
                                'checkout': reserva_data['fecha_salida']
                            },
                            'huespedes': [],
                            'habitaciones': [],
                            'precio_total': str(reserva_data.get('precio', 0)),
                            'nacionalidad': reserva_data.get('nacionalidad', '').upper()
                        }
                        
                        # Convertir pasajeros al formato de huespedes
                        for pasajero in reserva_data.get('pasajeros', []):
                            # Dividir nombre y apellido (asumiendo que el último espacio separa nombre de apellido)
                            partes = pasajero.split()
                            if len(partes) >= 2:
                                nombre = ' '.join(partes[:-1]).upper()  # Todo excepto el último como nombre
                                apellido = partes[-1].upper()  # Último como apellido
                            else:
                                nombre = pasajero.upper()
                                apellido = ''
                            
                            datos_adaptados['huespedes'].append({
                                'nombre': nombre,
                                'apellido': apellido
                            })
                        
                        # Convertir habitaciones al formato esperado
                        for hab in reserva_data.get('habitaciones', []):
                            datos_adaptados['habitaciones'].append({
                                'tipo': hab.get('tipo', '').upper(),
                                'adultos': hab.get('adultos', 0),
                                'ninios': hab.get('ninos', 0) + hab.get('bebes', 0),  # Children + Babies
                                'regimen': reserva_data.get('meal_plan', 'ALL INCLUSIVE').upper()
                            })
                        
                        # Usar la función existente para crear la reserva
                        if crear_reserva_desde_bedbankglobal(datos_adaptados, request.user):
                            reservas_creadas += 1
                            print(f"✅ Reserva creada exitosamente: {reserva_data['booking_code']}")
                            messages.success(request, f"Reserva {reserva_data['booking_code']} importada correctamente")
                        else:
                            reservas_con_error += 1
                            messages.error(request, f"Error al crear reserva {reserva_data['booking_code']}")
                            
                    except IntegrityError as e:
                        if 'UNIQUE constraint' in str(e):
                            print(f"❌ Error: Booking code duplicado {reserva_data.get('booking_code', 'N/A')}")
                            messages.error(request, f"Reserva {reserva_data.get('booking_code', 'N/A')}: Booking code ya existe")
                        else:
                            print(f"❌ Error de integridad: {e}")
                            messages.error(request, f"Error al crear reserva {reserva_data.get('booking_code', 'N/A')}: {e}")
                        reservas_con_error += 1
                        
                    except Exception as e:
                        print(f"❌ Error inesperado creando reserva {reserva_data.get('booking_code', 'N/A')}: {e}")
                        messages.error(request, f"Error inesperado con reserva {reserva_data.get('booking_code', 'N/A')}: {str(e)}")
                        reservas_con_error += 1
                
                # Resumen final
                if reservas_creadas > 0:
                    messages.success(request, f"✅ {reservas_creadas} reserva(s) importada(s) correctamente")
                if reservas_con_error > 0:
                    messages.warning(request, f"⚠️ {reservas_con_error} reserva(s) no se pudieron importar")
                
                return redirect('reservations:list')
                
            except Exception as e:
                print(f"❌ Error general procesando archivo: {e}")
                messages.error(request, f"Error procesando archivo: {str(e)}")
                return redirect('reservations:importar_bedbankglobal')
        
        else:
            messages.error(request, "No se seleccionó ningún archivo.")
            return redirect('reservations:importar_bedbankglobal')
    
    # GET request - mostrar formulario
    context = {
        'title': 'Importar BedBankGlobal',
        'active_menu': 'importar',
    }
    return render(request, 'reservations/importar_bedbankglobal.html', context)

def crear_reserva_desde_bedbankglobal(data_reserva, user):
    """Crear una reserva en la base de datos desde datos de BedBankGlobal - CON UPPER CASE"""
    try:
        # Verificar si ya existe una reserva con este booking_code
        if Reservation.objects.filter(booking_code=data_reserva['voucher']).exists():
            print(f"Reserva con voucher {data_reserva['voucher']} ya existe")
            return False
        
        # Concatenar nombres de huéspedes - CONVERTIR A UPPER CASE
        nombres_clientes = ", ".join([
            f"{huesped.get('nombre', '')} {huesped.get('apellido', '')}".strip()
            for huesped in data_reserva['huespedes']
            if huesped.get('nombre') or huesped.get('apellido')
        ]).upper()  # CONVERTIR A UPPER CASE
        
        # Si no hay nombres, usar un valor por defecto
        if not nombres_clientes:
            nombres_clientes = "HUÉSPED NO ESPECIFICADO"  # EN UPPER CASE
        
        # Usar el régimen de la primera habitación o valor por defecto - CONVERTIR A UPPER CASE
        regimen = data_reserva['habitaciones'][0]['regimen'] if data_reserva['habitaciones'] else 'ALL INCLUSIVE'
        
        # Calcular precio total
        try:
            precio_venta = float(data_reserva['precio_total'].replace(',', '.')) if data_reserva['precio_total'] and data_reserva['precio_total'] != '0.00' else 0
        except (ValueError, AttributeError):
            precio_venta = 0
        
        # Usar nacionalidad del primer huésped o valor por defecto - CONVERTIR A UPPER CASE
        nacionalidad = data_reserva.get('nacionalidad', 'NO ESPECIFICADA')
        
        # Crear la reserva principal - TODOS LOS CAMPOS DE TEXTO EN UPPER CASE
        reserva = Reservation(
            status='OK',
            agency='BEDBANKGLOBAL',  # UPPER CASE
            booking_code=data_reserva['voucher'],
            clients_names=nombres_clientes,  # YA EN UPPER CASE
            hotel=data_reserva['hotel'],  # YA EN UPPER CASE
            date_from=data_reserva['fechas']['checkin'],
            date_to=data_reserva['fechas']['checkout'],
            meal_plan=regimen,  # YA EN UPPER CASE
            sale_price=precio_venta,
            touch_cost=0,
            nationality=nacionalidad,  # YA EN UPPER CASE
            remarks=f"IMPORTADO AUTOMÁTICAMENTE DESDE BEDBANKGLOBAL. {len(data_reserva['habitaciones'])} HABITACIÓN(ES).".upper(),  # UPPER CASE
            created_by=user
        )
        reserva.save()
        
        # Crear todas las habitaciones - TIPO DE HABITACIÓN EN UPPER CASE
        for hab_data in data_reserva['habitaciones']:
            room = Room(
                reservation=reserva,
                room_type=hab_data['tipo'],  # YA EN UPPER CASE
                pax_ad=hab_data['adultos'],
                pax_chd=hab_data['ninios']  # SUMA DE CHILDREN + BABIES
            )
            room.save()
        
        print(f"Reserva BedBankGlobal creada exitosamente: {reserva.booking_code}")
        return True
        
    except Exception as e:
        print(f"Error creando reserva BedBankGlobal: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return False